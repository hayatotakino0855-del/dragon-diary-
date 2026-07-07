"""基準Excelから特性定義JSON(samples/ko5155b.json)を抽出生成する。

手転記による文字ずれ(全角スペース等)を避けるため、名称・定型文は
基準Excelのセル値をそのまま使う。Step1受け入れテストの入力にもなる。
"""
import json
import os
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from app.engine import prototypes as P  # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))


def build():
    wb = load_workbook(P.REFERENCE_PATH)
    g = wb[P.SHEET_GAIRIN]
    n = wb[P.SHEET_NAIRIN]

    def common(ws, ring_cell="B4", date_cell="K4"):
        return dict(
            ring_type=ws[ring_cell].value,
            part_no=ws["C6"].value,
            material=ws["J6"].value,
            drawing_no=ws["C7"].value.split("：", 1)[1],
            container_qty_note=ws["J3"].value.split("：　", 1)[1],
            frequency_note=ws["B11"].value[len(P.FREQUENCY_PREFIX):],
            date_label=ws[date_cell].value,
        )

    gairin = dict(
        name="外輪", sheet_name=P.SHEET_GAIRIN, **common(g),
        special_notes=g["A337"].value,
        characteristics=[
            dict(number=1, type="bilateral_with_sub", name=g["B13"].value,
                 center_value="22.8", tolerance="0.3",
                 sub_name=g["B39"].value, sub_tolerance="0.3", axis_range=1.2),
            dict(number=2, type="bilateral", name=g["B65"].value,
                 center_value="84.3", tolerance="0.3"),
            dict(number=3, type="bilateral", name=g["B109"].value,
                 center_value="84.1", tolerance="0.3"),
            dict(number=4, type="bilateral", name=g["B153"].value,
                 center_value="62.2", tolerance="0.3"),
            dict(number=5, type="bilateral", name=g["B217"].value,
                 center_value=None, tolerance=None),
            dict(number=5, type="unilateral", name=g["B241"].value,
                 center_value="0.8", tolerance="0.8"),
            dict(number=6, type="unilateral", name=g["B264"].value,
                 center_value="0.8", tolerance="0.8"),
            dict(number=7, type="weight", name=g["B287"].value,
                 center_value="445.0", tolerance="4.0"),
            dict(number=8, type="visual", name=g["B332"].value,
                 criteria=g["B333"].value),
        ],
    )

    nairin = dict(
        name="内輪", sheet_name=P.SHEET_NAIRIN, **common(n),
        special_notes=n["A304"].value,
        characteristics=[
            dict(number=5, type="measured", name=n["B13"].value),
            dict(number=1, type="bilateral_with_sub", name=n["B24"].value,
                 center_value="22.9", tolerance="0.3",
                 sub_name=n["B50"].value, sub_tolerance="0.3", axis_range=1.2),
            dict(number=2, type="bilateral", name=n["B76"].value,
                 center_value="63.2", tolerance="0.3"),
            dict(number=3, type="bilateral", name=n["B120"].value,
                 center_value="62.3", tolerance="0.3"),
            dict(number=4, type="bilateral", name=n["B164"].value,
                 center_value="41.7", tolerance="0.3"),
            dict(number=5, type="unilateral", name=n["B208"].value,
                 center_value="0.8", tolerance="0.8"),
            dict(number=6, type="unilateral", name=n["B231"].value,
                 center_value="0.8", tolerance="0.8"),
            dict(number=7, type="weight", name=n["B254"].value,
                 center_value="302.0", tolerance="4.0"),
            dict(number=8, type="visual", name=n["B299"].value,
                 criteria=n["B300"].value),
        ],
    )
    return [gairin, nairin]


if __name__ == "__main__":
    data = build()
    out = os.path.join(HERE, "ko5155b.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print("written:", out)
