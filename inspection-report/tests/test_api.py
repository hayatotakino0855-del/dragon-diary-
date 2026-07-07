"""Web API のエンドツーエンドテスト。

一時DBを使い、サンプル投入 → 出力Excelが基準Excelと一致することまで確認する。
"""
import os
import sys
import tempfile

os.environ["INSPECTION_DB"] = os.path.join(tempfile.mkdtemp(), "test.db")

import pytest  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from app.main import app  # noqa: E402

client = TestClient(app)


@pytest.fixture(scope="module", autouse=True)
def seeded():
    r = client.post("/api/seed")
    assert r.status_code == 200 and r.json()["ok"]


def _template_ids():
    rows = client.get("/api/tree").json()
    return [r["template_id"] for r in rows if r["template_id"]]


def test_tree_and_filter():
    rows = client.get("/api/tree").json()
    assert len([r for r in rows if r["template_id"]]) == 2
    rows = client.get("/api/tree", params={"drawing_no": "KO-5155B"}).json()
    assert rows
    rows = client.get("/api/tree", params={"drawing_no": "存在しない"}).json()
    assert not [r for r in rows if r["product_id"]]


def test_get_update_template():
    tid = _template_ids()[0]
    t = client.get(f"/api/templates/{tid}").json()
    assert t["characteristics"]
    r = client.put(f"/api/templates/{tid}", json={
        "name": t["name"], "sheet_name": t["sheet_name"],
        "special_notes": t["special_notes"], "date_label": t["date_label"],
        "author": "テスト太郎",
        "characteristics": [
            {k: c[k] for k in ("number", "name", "type", "center_value",
                               "tolerance", "unit", "sub_name", "sub_tolerance",
                               "criteria", "axis_step", "axis_range")}
            for c in t["characteristics"]],
    })
    assert r.status_code == 200
    assert r.json()["warnings"] == []
    assert client.get(f"/api/templates/{tid}").json()["author"] == "テスト太郎"


def test_export_matches_reference(tmp_path):
    """API経由の出力も基準Excelと一致する(Step1テストと同じ検証)。"""
    from openpyxl import load_workbook

    from app.engine import prototypes as P
    from tests.test_regenerate import (KNOWN_REFERENCE_INCONSISTENCIES,
                                       _cell_sig)

    ids = ",".join(map(str, _template_ids()))
    r = client.get(f"/api/export?ids={ids}")
    assert r.status_code == 200
    assert "KO-5155B-HT_6212RSH2" in r.headers["content-disposition"]
    out = tmp_path / "api_export.xlsx"
    out.write_bytes(r.content)
    gen_wb = load_workbook(out)
    ref_wb = load_workbook(P.REFERENCE_PATH)
    for sheet in ref_wb.sheetnames:
        ref, gen = ref_wb[sheet], gen_wb[sheet]
        allowed = KNOWN_REFERENCE_INCONSISTENCIES.get(sheet, set())
        for row in ref.iter_rows(max_col=21):
            for c in row:
                if c.coordinate in allowed:
                    continue
                assert _cell_sig(gen[c.coordinate]) == _cell_sig(c), c.coordinate


def test_duplicate_and_revision():
    tid = _template_ids()[0]
    r = client.post(f"/api/templates/{tid}/duplicate",
                    json={"as_revision": False}).json()
    dup = client.get(f"/api/templates/{r['id']}").json()
    assert "複製" in dup["name"] and dup["revision"] == 1
    client.delete(f"/api/templates/{r['id']}")

    r = client.post(f"/api/templates/{tid}/duplicate",
                    json={"as_revision": True}).json()
    new = client.get(f"/api/templates/{r['id']}").json()
    old = client.get(f"/api/templates/{tid}").json()
    assert new["revision"] == old["revision"] + 1
    assert old["archived"] == 1 and new["archived"] == 0
    assert len(new["characteristics"]) == len(old["characteristics"])


def test_preview_html():
    tid = _template_ids()[1]
    r = client.get(f"/api/templates/{tid}/preview")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(("text/html", "application/pdf"))


def test_char_master_crud():
    r = client.post("/api/char-masters", json={
        "name": "外径", "type": "bilateral", "tolerance": "0.3"})
    mid = r.json()["id"]
    assert any(m["id"] == mid for m in client.get("/api/char-masters").json())
    client.delete(f"/api/char-masters/{mid}")
    assert not any(m["id"] == mid for m in client.get("/api/char-masters").json())


def test_fit_warning_and_auto_shrink():
    rows = client.get("/api/tree").json()
    pid = rows[0]["product_id"]
    chars = [dict(name=f"特性{i}", type="bilateral", center_value="10.0",
                  tolerance="0.3", axis_range=1.0) for i in range(9)]
    r = client.post(f"/api/products/{pid}/templates", json={
        "name": "詰め込み", "characteristics": chars}).json()
    assert r["warnings"], "9特性×軸1.0 は1ページ超過の警告が出るはず"
    s = client.post(f"/api/templates/{r['id']}/auto-shrink").json()
    assert s["changed"]
    client.delete(f"/api/templates/{r['id']}")
