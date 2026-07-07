"""Step1 受け入れテスト:

基準Excelと同じ特性定義を与えたとき、セル値・結合セル・列幅・行高・
罫線・フォント・塗り・配置・表示形式が基準と一致すること。
"""
import os
import sys

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from app.engine import prototypes as P  # noqa: E402
from app.engine.generator import generate_workbook  # noqa: E402
from samples.make_samples import build  # noqa: E402

MAX_COL = 21

# 基準Excel自体の内部不整合(外輪と内輪で書式が食い違う箇所)。
# 生成器は外輪シートの書式を正として統一するため、内輪側の以下のセルは
# 書式比較から除外する(値の一致は要求する)。
#   A12:C12 … ヘッダ「特性」ラベルの太字/配置が外輪のみ設定されている
#   B139/B183 … 中心値の表示形式が内輪の2箇所だけ '0.0'(他は 'General'。
#                いずれも小数1桁の値のため印刷結果は同一)
KNOWN_REFERENCE_INCONSISTENCIES = {
    P.SHEET_NAIRIN: {"A12", "B12", "C12", "B139", "B183"},
}


def _cell_sig(c):
    f, a, fl = c.font, c.alignment, c.fill
    return dict(
        value=c.value,
        font=(f.name, f.size, f.bold, f.italic, f.underline,
              getattr(f.color, "rgb", None) if f.color else None),
        align=(a.horizontal, a.vertical, a.wrap_text, a.shrink_to_fit,
               a.text_rotation, a.indent),
        fill=(fl.patternType,
              getattr(fl.fgColor, "rgb", None) if fl.fgColor else None),
        numfmt=c.number_format,
    )


def _edge_borders(ws, max_row):
    """印刷上の実効罫線をエッジ単位で取り出す。

    隣接セルのどちら側に罫線が格納されていても同一の線として扱う
    (基準Excel自体にブロック間で格納側が異なる箇所があるため)。
    """
    edges = {}
    for r in range(1, max_row + 1):
        for c in range(1, MAX_COL + 1):
            b = ws.cell(row=r, column=c).border
            below = ws.cell(row=r + 1, column=c).border if r < max_row else None
            right = ws.cell(row=r, column=c + 1).border if c < MAX_COL else None
            h = b.bottom.style or (below.top.style if below else None)
            v = b.right.style or (right.left.style if right else None)
            if h:
                edges[("h", r, c)] = h
            if v:
                edges[("v", r, c)] = v
            if r == 1 and b.top.style:
                edges[("h", 0, c)] = b.top.style
            if c == 1 and b.left.style:
                edges[("v", r, 0)] = b.left.style
    return edges


def _col_widths(ws):
    widths = {}
    for dim in ws.column_dimensions.values():
        if dim.width is not None and dim.min and dim.max:
            for i in range(dim.min, min(dim.max, MAX_COL) + 1):
                widths[get_column_letter(i)] = round(dim.width, 4)
    return widths


@pytest.fixture(scope="module")
def generated(tmp_path_factory):
    out = tmp_path_factory.mktemp("out") / "regen.xlsx"
    wb = generate_workbook(build())
    wb.save(out)
    return load_workbook(out)


@pytest.fixture(scope="module")
def reference():
    return load_workbook(P.REFERENCE_PATH)


@pytest.mark.parametrize("sheet", [P.SHEET_GAIRIN, P.SHEET_NAIRIN])
def test_sheet_matches_reference(generated, reference, sheet):
    ref = reference[sheet]
    assert sheet in generated.sheetnames, f"シート {sheet!r} が生成されていない"
    gen = generated[sheet]

    assert gen.max_row >= ref.max_row

    # 結合セル
    ref_merges = {str(m) for m in ref.merged_cells.ranges if m.min_col <= MAX_COL}
    gen_merges = {str(m) for m in gen.merged_cells.ranges if m.min_col <= MAX_COL}
    assert gen_merges == ref_merges, (
        f"結合セル不一致 missing={sorted(ref_merges - gen_merges)[:10]} "
        f"extra={sorted(gen_merges - ref_merges)[:10]}")

    # 行高
    for r in range(1, ref.max_row + 1):
        rh = ref.row_dimensions[r].height
        gh = gen.row_dimensions[r].height
        assert (rh or 13.0) == pytest.approx(gh or 13.0), f"行高不一致 row={r} {rh} != {gh}"

    # 列幅
    assert _col_widths(gen) == _col_widths(ref), "列幅不一致"

    # セル値と書式
    allowed = KNOWN_REFERENCE_INCONSISTENCIES.get(sheet, set())
    diffs = []
    for r in range(1, ref.max_row + 1):
        for c in range(1, MAX_COL + 1):
            rs = _cell_sig(ref.cell(row=r, column=c))
            gs = _cell_sig(gen.cell(row=r, column=c))
            if f"{get_column_letter(c)}{r}" in allowed:
                assert rs["value"] == gs["value"], f"許容セルでも値は一致必須: {get_column_letter(c)}{r}"
                continue
            if rs != gs:
                keys = [k for k in rs if rs[k] != gs[k]]
                diffs.append(f"{get_column_letter(c)}{r}: {keys} ref={[rs[k] for k in keys]} gen={[gs[k] for k in keys]}")
    assert not diffs, "セル不一致 ({}件):\n".format(len(diffs)) + "\n".join(diffs[:30])

    # 罫線(エッジ単位の実効比較)
    ref_edges = _edge_borders(ref, ref.max_row)
    gen_edges = _edge_borders(gen, ref.max_row)
    def _edge_allowed(k):
        kind, r, c = k
        cells = [(r, c), (r + 1, c)] if kind == "h" else [(r, c), (r, c + 1)]
        return any(f"{get_column_letter(cc)}{rr}" in allowed
                   for rr, cc in cells if rr >= 1 and cc >= 1)

    edge_diffs = [(k, ref_edges.get(k), gen_edges.get(k))
                  for k in set(ref_edges) | set(gen_edges)
                  if ref_edges.get(k) != gen_edges.get(k) and not _edge_allowed(k)]
    assert not edge_diffs, f"罫線不一致 ({len(edge_diffs)}件): {sorted(edge_diffs)[:20]}"

    # 印刷設定
    assert gen.page_setup.paperSize == ref.page_setup.paperSize
    assert gen.page_setup.orientation == ref.page_setup.orientation
    for attr in ("left", "right", "top", "bottom"):
        assert getattr(gen.page_margins, attr) == pytest.approx(
            getattr(ref.page_margins, attr))
