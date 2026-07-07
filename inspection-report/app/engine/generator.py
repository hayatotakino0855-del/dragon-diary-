"""帳票生成エンジン。

「基準Excelをテンプレートとして複製し、ブロック単位で組み替える」方式。
基準ワークブックを読み込み、その中に新しいシートを組み立ててから
元シートを削除する。同一ワークブック内でスタイルを共有コピーするため
罫線・フォントの再現ずれが起きない。
"""
from decimal import Decimal

from openpyxl import load_workbook

from . import prototypes as P
from .axis import bilateral_labels, unilateral_labels
from .rowcopy import copy_row_range, copy_single_row_styles, copy_sheet_settings

CONTENT_START_ROW = 13  # ヘッダは行1..12


# ---------------------------------------------------------------- utilities
def _num_or_str(v):
    """基準Excelは数値セル(84.3 等)は数値型で持つ。同じ型で書き込む。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    try:
        f = float(s)
        return int(f) if f == int(f) and "." not in s else f
    except ValueError:
        return s


def _dec(v, default):
    return Decimal(str(v if v is not None else default))


def _axis_params(char, proto):
    rng = char.get("axis_range") or proto.get("axis_range")
    step = char.get("axis_step") or proto.get("axis_step")
    return Decimal(str(rng)), Decimal(str(step))


def _resolve_type(char):
    t = char["type"]
    if t == "bilateral" and char.get("center_value") in (None, ""):
        return "bilateral_plain"
    return t


# ---------------------------------------------------------------- blocks
def _write_number(ws, row, number):
    ws.cell(row=row, column=1).value = number


def _set(ws, row, col, value):
    ws.cell(row=row, column=col).value = value


def _render_block(wb, ws, dst_row, char, seq_number):
    """1特性ブロックを dst_row から描画し、使用行数を返す。"""
    tname = _resolve_type(char)
    proto = P.BLOCKS[tname]
    src_ws = wb[proto["sheet"]]
    number = char.get("number", seq_number)

    if tname in ("weight", "visual", "measured"):
        rows = copy_row_range(src_ws, proto["start"], proto["end"], ws, dst_row)
        _write_number(ws, dst_row, number)
        _set(ws, dst_row + proto["name_cell"], 2, char["name"])
        if tname == "weight":
            _set(ws, dst_row + proto["center_cell"], 2, f"{char['center_value']} gr")
            _set(ws, dst_row + proto["tol_cell"], 2, f"±{char['tolerance']} gr")
        elif tname == "visual":
            _set(ws, dst_row + proto["criteria_cell"], 2,
                 char.get("criteria") or "有害な欠陥無き事")
        return rows

    rng, step = _axis_params(char, proto)
    verbatim = (rng == Decimal(str(proto["axis_range"]))
                and step == Decimal(str(proto["axis_step"])))

    if tname in ("bilateral", "bilateral_plain", "bilateral_with_sub"):
        if verbatim:
            rows = copy_row_range(src_ws, proto["start"], proto["end"], ws, dst_row)
        else:
            rows = _synth_bilateral(src_ws, proto, ws, dst_row, rng, step,
                                    with_merges=proto["sections"] is not None)
        _write_number(ws, dst_row, number)
        _substitute_bilateral_values(ws, dst_row, char, tname, proto, rng, step,
                                     verbatim)
        return rows

    if tname == "unilateral":
        if verbatim:
            rows = copy_row_range(src_ws, proto["start"], proto["end"], ws, dst_row)
            name_r, val_r, ika_r = (dst_row + proto["name_cell"],
                                    dst_row + proto["center_cell"],
                                    dst_row + proto["tol_cell"])
        else:
            rows, name_r, val_r, ika_r = _synth_unilateral(
                src_ws, proto, ws, dst_row, rng, step)
        _write_number(ws, dst_row, number)
        _set(ws, name_r, 2, char["name"])
        _set(ws, val_r, 2, _num_or_str(char["center_value"] or char["tolerance"]))
        _set(ws, ika_r, 2, "以下")
        return rows

    raise ValueError(f"unknown characteristic type: {tname}")


def _substitute_bilateral_values(ws, dst_row, char, tname, proto, rng, step,
                                 verbatim):
    if tname == "bilateral_plain":
        if verbatim:
            name_row = dst_row + proto["name_cell"]
        else:
            u = int(rng / step)  # 片側目盛り数
            name_row = dst_row + 1 + 2 * (u - 1) + 1  # 「0」目盛りペアの直前行
        _set(ws, name_row, 2, char["name"])
        return
    if verbatim:
        name_row = dst_row + proto["name_cell"]
        center_row = dst_row + proto["center_cell"]
    else:
        u = int(rng / step)
        c_abs = dst_row + 1 + 2 * (u - 1)  # 「+s」目盛りペア先頭行
        name_row = dst_row
        center_row = c_abs if tname == "bilateral" else c_abs - 1
    _set(ws, name_row, 2, char["name"])
    if tname == "bilateral":
        _set(ws, center_row, 2, _num_or_str(char["center_value"]))
        tol_row = (dst_row + proto["tol_cell"]) if verbatim else center_row + 5
        _set(ws, tol_row, 2, f"±{char['tolerance']}")
    else:  # bilateral_with_sub
        _set(ws, center_row, 2, f"{char['center_value']}±{char['tolerance']}")
        sub_name_row = (dst_row + proto["sub_name_cell"]) if verbatim else center_row + 4
        sub_tol_row = (dst_row + proto["sub_tol_cell"]) if verbatim else center_row + 8
        _set(ws, sub_name_row, 2, char.get("sub_name") or "")
        _set(ws, sub_tol_row, 2, f"{char.get('sub_tolerance')} 以下")


def _synth_bilateral(src_ws, proto, ws, dst_row, rng, step, with_merges):
    """軸範囲/刻みが雛形と異なる両側公差ブロックを合成する。

    雛形の行スタイルを中心から外へ位置対応でコピーし、延長分は最外周の
    目盛りペアのスタイルを繰り返す。
    """
    start = proto["start"]
    upper, center3, lower = bilateral_labels(rng, step)
    u = len(upper)  # 中心3目盛りを除く片側目盛り数
    u0 = len(proto["upper_ticks"])

    plan = []  # (src_row, label or None, is_pair_head)
    plan.append((start + proto["pad_top"], None))
    for j in range(u - 1, -1, -1):  # j = 中心からの距離(0が内側)
        off = proto["upper_ticks"][u0 - 1 - j] if j < u0 else proto["upper_ticks"][0]
        plan.append((start + off, upper[u - 1 - j]))
        plan.append((start + off + 1, None))
    for k, off in enumerate(proto["center_ticks"]):
        plan.append((start + off, center3[k]))
        plan.append((start + off + 1, None))
    for j in range(0, u):
        off = proto["lower_ticks"][j] if j < u0 else proto["lower_ticks"][-1]
        plan.append((start + off, lower[j]))
        plan.append((start + off + 1, None))
    plan.append((start + proto["pad_bottom"], None))

    d = dst_row
    tick_rows = []
    for src_row, label in plan:
        copy_single_row_styles(src_ws, src_row, ws, d)
        if label is not None:
            ws.cell(row=d, column=4).value = label
            tick_rows.append(d)
        d += 1
    total = d - dst_row
    end = dst_row + total - 1

    for r in tick_rows:
        ws.merge_cells(start_row=r, start_column=4, end_row=r + 1, end_column=4)
    ws.merge_cells(start_row=dst_row, start_column=1, end_row=end, end_column=1)
    if with_merges:
        c_abs = dst_row + 1 + 2 * u  # 「+s」目盛りペア先頭行
        for s_row, e_row in proto["sections"](dst_row, c_abs, end):
            ws.merge_cells(start_row=s_row, start_column=2, end_row=e_row, end_column=3)
    return total


def _synth_unilateral(src_ws, proto, ws, dst_row, rng, step):
    start = proto["start"]
    labels = unilateral_labels(rng, step)
    n = len(labels)
    n0 = len(proto["ticks"])

    plan = [(start + proto["pad_top"], None)]
    for i in range(n):
        off = proto["ticks"][i] if i < n0 else proto["ticks"][-1]
        plan.append((start + off, labels[i]))
        plan.append((start + off + 1, None))

    d = dst_row
    tick_rows = []
    for src_row, label in plan:
        copy_single_row_styles(src_ws, src_row, ws, d)
        if label is not None:
            ws.cell(row=d, column=4).value = label
            tick_rows.append(d)
        d += 1
    total = d - dst_row
    end = dst_row + total - 1

    for r in tick_rows[:-1]:
        ws.merge_cells(start_row=r, start_column=4, end_row=r + 1, end_column=4)
    # 最下段「0」も基準どおり2行結合
    ws.merge_cells(start_row=tick_rows[-1], start_column=4,
                   end_row=tick_rows[-1] + 1, end_column=4)
    ws.merge_cells(start_row=dst_row, start_column=1, end_row=end, end_column=1)

    f1, f2 = proto["section_fracs"]
    name_len = round(total * f1)
    val_len = round(total * f2)
    name_end = dst_row + name_len - 1
    val_end = name_end + val_len
    ws.merge_cells(start_row=dst_row, start_column=2, end_row=name_end, end_column=3)
    ws.merge_cells(start_row=name_end + 1, start_column=2, end_row=val_end, end_column=3)
    ws.merge_cells(start_row=val_end + 1, start_column=2, end_row=end, end_column=3)
    return total, dst_row, name_end + 1, val_end + 1


# ---------------------------------------------------------------- sheet
def build_sheet(wb, template, index=None):
    """テンプレート定義 1 件から 1 シートを組み立てる。"""
    title = template.get("sheet_name") or template["name"]
    ws = wb.create_sheet(title=title, index=index)
    header_ws = wb[P.HEADER["sheet"]]
    copy_sheet_settings(header_ws, ws)
    copy_row_range(header_ws, P.HEADER["start"], P.HEADER["end"], ws, 1)

    f = P.HEADER_FIELDS
    _set(ws, *f["ring_type"], template.get("ring_type") or template["name"])
    _set(ws, *f["part_no"], template.get("part_no") or "")
    _set(ws, *f["material"], template.get("material") or "")
    _set(ws, *f["drawing_no"], P.DRAWING_NO_PREFIX + (template.get("drawing_no") or ""))
    _set(ws, *f["container_qty"],
         P.CONTAINER_QTY_PREFIX + (template.get("container_qty_note") or ""))
    _set(ws, *f["frequency"], P.FREQUENCY_PREFIX + (template.get("frequency_note") or ""))
    _set(ws, *f["date_label"], template.get("date_label") or P.DEFAULT_DATE_LABEL)

    row = CONTENT_START_ROW
    for i, char in enumerate(template["characteristics"]):
        row += _render_block(wb, ws, row, char, i + 1)

    footer_ws = wb[P.FOOTER["sheet"]]
    copy_row_range(footer_ws, P.FOOTER["start"], P.FOOTER["end"], ws, row)
    notes = template.get("special_notes")
    if notes:
        nr, nc = P.FOOTER_NOTES_CELL
        _set(ws, row + (nr - P.FOOTER["start"]), nc, notes)
    return ws


def generate_workbook(templates, reference_path=None):
    """テンプレート定義のリストから 1 ワークブック(1定義=1シート)を生成する。"""
    wb = load_workbook(reference_path or P.REFERENCE_PATH)
    original = list(wb.sheetnames)
    built = []
    for t in templates:
        ws = build_sheet(wb, t)
        built.append((ws, t.get("sheet_name") or t["name"]))
    for name in original:
        del wb[name]
    for ws, title in built:  # 基準シートと同名の場合の連番回避を解消
        ws.title = title
    return wb


# ---------------------------------------------------------------- page fit
def _block_height(wb, char):
    tname = _resolve_type(char)
    proto = P.BLOCKS[tname]
    src_ws = wb[proto["sheet"]]
    if tname in ("weight", "visual", "measured"):
        return _range_height(src_ws, proto["start"], proto["end"])
    rng, step = _axis_params(char, proto)
    if (rng == Decimal(str(proto["axis_range"]))
            and step == Decimal(str(proto["axis_step"]))):
        return _range_height(src_ws, proto["start"], proto["end"])
    if tname == "unilateral":
        n = int(rng / step) + 1
        return (1 + 2 * n) * 3.65
    n = int(rng / step)  # 片側(符号付き含む)
    return (2 + 2 * (2 * n + 1)) * 3.65


def _range_height(ws, start, end):
    return sum((ws.row_dimensions[r].height or 13.0) for r in range(start, end + 1))


def fit_report(templates, reference_path=None):
    """各テンプレートの高さと1ページ予算(基準Excel実績)を返す。"""
    wb = load_workbook(reference_path or P.REFERENCE_PATH)
    budget = _range_height(wb[P.SHEET_GAIRIN], 1, 341)
    report = []
    for t in templates:
        h = _range_height(wb[P.HEADER["sheet"]], P.HEADER["start"], P.HEADER["end"])
        h += _range_height(wb[P.FOOTER["sheet"]], P.FOOTER["start"], P.FOOTER["end"])
        for c in t["characteristics"]:
            h += _block_height(wb, c)
        report.append(dict(name=t.get("sheet_name") or t["name"],
                           height=round(h, 2), budget=round(budget, 2),
                           fits=h <= budget + 0.01))
    return report


def auto_shrink(template, reference_path=None):
    """収まらない場合に軸延長範囲を段階的に短縮する(仕様の第一手段)。

    公差+2目盛りを下限として、最も長い軸から1刻みずつ縮める。
    戻り値: 変更があれば True。
    """
    changed = False
    while True:
        rep = fit_report([template], reference_path)[0]
        if rep["fits"]:
            return changed
        cands = []
        for c in template["characteristics"]:
            t = _resolve_type(c)
            if t not in ("bilateral", "bilateral_plain", "bilateral_with_sub",
                         "unilateral"):
                continue
            proto = P.BLOCKS[t]
            rng, step = _axis_params(c, proto)
            try:
                tol = Decimal(str(c.get("tolerance") or 0))
            except Exception:
                tol = Decimal(0)
            floor = tol + 2 * step
            if rng - step >= floor:
                cands.append((rng, step, c))
        if not cands:
            return changed
        cands.sort(key=lambda x: x[0], reverse=True)
        rng, step, c = cands[0]
        c["axis_range"] = float(rng - step)
        changed = True
