"""LibreOffice が無い環境向けの HTML 簡易プレビュー。

罫線・結合・行高を近似再現する。印刷イメージの確認用であり、
正式な出力は Excel ファイル側。
"""
import html

MAX_COL = 21
BORDER_CSS = {
    "hair": "0.5px solid #999",
    "thin": "1px solid #333",
    "medium": "2px solid #000",
    "thick": "3px solid #000",
    "double": "3px double #000",
    "dotted": "1px dotted #666",
    "dashed": "1px dashed #666",
}


def _b(style):
    return BORDER_CSS.get(style, "none") if style else "none"


def worksheet_to_html(ws, warnings=()):
    merged = {}
    covered = set()
    for m in ws.merged_cells.ranges:
        merged[(m.min_row, m.min_col)] = (m.max_row - m.min_row + 1,
                                          m.max_col - m.min_col + 1)
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                if (r, c) != (m.min_row, m.min_col):
                    covered.add((r, c))

    from openpyxl.utils import get_column_letter
    widths = []
    for c in range(1, MAX_COL + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        w = (dim.width if dim and dim.width else 8.4)
        widths.append(round(w * 7))

    out = ["<table style='border-collapse:collapse;table-layout:fixed;"
           "font-family:\"MS PGothic\",\"Yu Gothic\",sans-serif'>"]
    out.append("<colgroup>" + "".join(
        f"<col style='width:{w}px'>" for w in widths) + "</colgroup>")
    for r in range(1, ws.max_row + 1):
        h = ws.row_dimensions[r].height or 13.0
        out.append(f"<tr style='height:{round(h * 1.33, 1)}px'>")
        for c in range(1, MAX_COL + 1):
            if (r, c) in covered:
                continue
            cell = ws.cell(row=r, column=c)
            span = merged.get((r, c))
            attrs = ""
            if span:
                attrs = f" rowspan={span[0]} colspan={span[1]}"
            bd = cell.border
            f = cell.font
            a = cell.alignment
            css = [
                f"border-left:{_b(bd.left.style)}",
                f"border-right:{_b(bd.right.style)}",
                f"border-top:{_b(bd.top.style)}",
                f"border-bottom:{_b(bd.bottom.style)}",
                f"font-size:{round((f.size or 11) * 1.1, 1)}px",
                "padding:0", "overflow:hidden", "line-height:1.05",
            ]
            if f.bold:
                css.append("font-weight:bold")
            css.append(f"text-align:{a.horizontal or 'left'}")
            css.append("vertical-align:" + {"center": "middle"}.get(
                a.vertical or "bottom", a.vertical or "bottom"))
            if a.wrap_text:
                css.append("white-space:pre-wrap")
            else:
                # Excelは非折返しセルの文字を隣の空セルへはみ出させる
                css.remove("overflow:hidden")
                css.append("white-space:nowrap")
                css.append("overflow:visible")
            v = cell.value
            text = html.escape(str(v)) if v is not None else ""
            text = text.replace("\n", "<br>")
            out.append(f"<td{attrs} style='{';'.join(css)}'>{text}</td>")
        out.append("</tr>")
    out.append("</table>")
    warn = "".join(f"<p style='color:#b00;font-weight:bold'>⚠ {html.escape(w)}</p>"
                   for w in warnings)
    return ("<!doctype html><meta charset='utf-8'>"
            f"<body style='background:#eee;padding:16px'>{warn}"
            f"<div style='background:#fff;padding:12px;width:fit-content'>"
            f"{''.join(out)}</div></body>")
